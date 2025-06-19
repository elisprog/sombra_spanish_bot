import os
import re
import asyncio
import nest_asyncio
from aiogram import Bot, Dispatcher, types, F
from aiogram.enums import ParseMode
from aiogram.client.default import DefaultBotProperties
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.filters import Command
from aiogram.types import FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery, ReplyKeyboardMarkup, KeyboardButton
from openpyxl import load_workbook
from moviepy.editor import VideoFileClip
import whisper
import difflib
import Levenshtein
import requests

# --- Клавиатура ---
start_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Обновить и перезапустить")],
        [KeyboardButton(text="Помощь")],
        [KeyboardButton(text="Выбрать фрагмент")],
        [KeyboardButton(text="Завершить на сегодня")]
    ],
    resize_keyboard=True,
    one_time_keyboard=False
)

# --- Видео ---
def download_video():
    file_id = "1qhkfTq6KujEBERAVBMjRGLnb-cX4FDEO"
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    path = "Extra1.mp4"
    if not os.path.exists(path):
        print("📥 Скачиваем видео...")
        session = requests.Session()
        response = session.get(url, stream=True)
        token = get_confirm_token(response)
        if token:
            url = f"https://drive.google.com/uc?export=download&confirm={token}&id={file_id}"
            response = session.get(url, stream=True)
        save_response_content(response, path)
        print("✅ Видео скачано.")

def get_confirm_token(response):
    for key, value in response.cookies.items():
        if key.startswith("download_warning"):
            return value
    return None

def save_response_content(response, destination):
    CHUNK_SIZE = 32768
    with open(destination, "wb") as f:
        for chunk in response.iter_content(CHUNK_SIZE):
            if chunk:
                f.write(chunk)

download_video()
nest_asyncio.apply()

# --- Настройки ---
TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "ВАШ_ТОКЕН")
VIDEO_PATH = "Extra1.mp4"
FRAGMENTS_DIR = "fragments"
USER_AUDIO_DIR = "user_audio"
EXCEL_PATH = "blocks.xlsx"
total = 23

os.makedirs(FRAGMENTS_DIR, exist_ok=True)
os.makedirs(USER_AUDIO_DIR, exist_ok=True)

keywords = [...]
bigrams = [...]

bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())
model = whisper.load_model("base")
user_state = {}

# --- Функции ---
def load_clips_from_excel(path=EXCEL_PATH):
    wb = load_workbook(path)
    ws = wb.active
    return [{"start": row[0], "end": row[1], "text": row[2]} for row in ws.iter_rows(min_row=2, values_only=True)]

clips = load_clips_from_excel()

def extract_clip(video_path: str, start: float, end: float, output_path: str):
    clip = VideoFileClip(video_path).subclip(start, end)
    clip.write_videofile(output_path, codec='libx264', audio_codec='aac', logger=None)

def generate_gap_text(sentence, keywords, bigrams):
    masked = sentence
    for bg in bigrams:
        masked = re.sub(rf"\b{re.escape(bg)}\b", "_____ _____", masked, flags=re.IGNORECASE)
    for kw in keywords:
        masked = re.sub(rf"\b{re.escape(kw)}\b", "_____", masked, flags=re.IGNORECASE)
    return masked

def transcribe(audio_path):
    result = model.transcribe(audio_path, language="es")
    return result["text"]

def normalize(text):
    text = text.lower()
    text = re.sub(r"[^\w\s]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def compare(reference, recognized):
    ref = normalize(reference)
    rec = normalize(recognized)
    distance = Levenshtein.distance(ref, rec)
    max_len = max(len(ref), len(rec))
    return 1 - distance / max_len if max_len > 0 else 0.0

def reset_user_fragment_state(chat_id):
    if chat_id in user_state:
        user_state[chat_id]["text_attempts"] = 0
        user_state[chat_id]["attempted_text"] = False
        user_state[chat_id]["audio_sent"] = False

def continue_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="✅ Продолжим?", callback_data="force_next")]
    ])

def navigation_keyboard():
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="🔙 Назад", callback_data="back"),
            InlineKeyboardButton(text="🔁 Повторить", callback_data="repeat"),
            InlineKeyboardButton(text="⏭ Пропустить", callback_data="skip")
        ]
    ])

def fragment_selection_keyboard():
    buttons, row = [], []
    for i in range(1, total + 1):
        row.append(InlineKeyboardButton(text=f"Фрагмент {i}", callback_data=f"goto_{i-1}"))
        if len(row) == 5:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    return InlineKeyboardMarkup(inline_keyboard=buttons)

# --- Команды и сообщения ---
@dp.message(F.text == "Обновить и перезапустить")
async def handle_start_text(message: types.Message):
    await cmd_start(message)

@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    user_state[message.chat.id] = {"index": 0, "attempted_text": False, "text_attempts": 0, "audio_sent": False}
    await message.answer("/start", reply_markup=start_keyboard)
    await send_clip(message.chat.id)

@dp.message(F.text == "Помощь")
@dp.message(Command("help"))
async def cmd_help(message: types.Message):
    await message.answer("ℹ️ <b>Помощь по боту</b>:\n/start — начать\n/select — выбрать фрагмент\n/help — помощь")

@dp.message(F.text == "Выбрать фрагмент")
@dp.message(Command("select"))
async def cmd_select(message: types.Message):
    await message.answer("Выбери номер фрагмента:", reply_markup=fragment_selection_keyboard())

@dp.message(F.text == "Завершить на сегодня")
async def handle_finish_today(message: types.Message):
    await message.answer("Спасибо за твою работу, ¡hasta luego!")

# --- Отправка фрагмента ---
async def send_clip(chat_id: int):
    state = user_state.get(chat_id)
    if not state or "index" not in state:
        await bot.send_message(chat_id, "⚠️ Пожалуйста, начни с команды /start.")
        return
    i = state["index"]
    if i >= len(clips):
        await bot.send_message(chat_id, "🎉 Все фрагменты завершены!")
        return
    reset_user_fragment_state(chat_id)
    clip = clips[i]
    clip_path = os.path.join(FRAGMENTS_DIR, f"clip_{i}.mp4")
    if not os.path.exists(clip_path):
        try:
            extract_clip(VIDEO_PATH, clip['start'], clip['end'], clip_path)
        except Exception as e:
            await bot.send_message(chat_id, f"❌ Ошибка: {e}")
            return
    await bot.send_video(chat_id, FSInputFile(clip_path))
    await bot.send_message(chat_id, f"🎬 <b>Фрагмент {i+1} из {total}</b>\n\n📝 Вставь пропущенные слова:\n\n{generate_gap_text(clip['text'], keywords, bigrams)}", reply_markup=navigation_keyboard())

# --- Callback ---
@dp.callback_query(lambda c: c.data in ["repeat", "skip", "back"])
async def handle_navigation_buttons(callback: CallbackQuery):
    chat_id = callback.message.chat.id
    state = user_state.get(chat_id)
    if not state:
        await callback.answer("Начни с /start")
        return
    if callback.data == "repeat":
        await callback.message.answer("🔁 Повторим:")
    elif callback.data == "skip":
        state["index"] += 1
        await callback.message.answer("⏭ Следующий фрагмент:")
    elif callback.data == "back":
        state["index"] = max(0, state["index"] - 1)
        await callback.message.answer("🔙 Назад:")
    reset_user_fragment_state(chat_id)
    await send_clip(chat_id)
    await callback.answer()

@dp.callback_query(lambda c: c.data.startswith("goto_"))
async def handle_goto_fragment(callback: CallbackQuery):
    chat_id = callback.message.chat.id
    index = int(callback.data.split("_")[1])
    if index < 0 or index >= len(clips):
        await callback.answer("Неверный номер")
        return
    user_state[chat_id]["index"] = index
    reset_user_fragment_state(chat_id)
    await callback.message.answer(f"Фрагмент {index + 1}")
    await send_clip(chat_id)
    await callback.answer()

@dp.callback_query(lambda c: c.data == "force_next")
async def handle_force_next(callback: CallbackQuery):
    chat_id = callback.message.chat.id
    state = user_state.get(chat_id)
    if not state:
        await callback.answer("Начни с /start")
        return
    if not state.get("audio_sent", False):
        await callback.message.answer("🎤 Пришли аудио с этим фрагментом.")
    else:
        state["index"] += 1
        reset_user_fragment_state(chat_id)
        await callback.message.answer("⏭ Следующий фрагмент.")
        await send_clip(chat_id)
    await callback.answer()

@dp.callback_query(lambda c: c.data == "show_answer")
async def show_answer_handler(callback: CallbackQuery):
    chat_id = callback.message.chat.id
    i = user_state[chat_id]["index"]
    await callback.message.answer(f"✅ Ответ:\n\n<b>{clips[i]['text']}</b>", reply_markup=continue_keyboard())
    await callback.answer()

# --- Сообщения от пользователя ---
@dp.message(F.text)
async def handle_text_response(message: types.Message):
    chat_id = message.chat.id
    state = user_state.get(chat_id)
    if not state:
        return
    i = state["index"]
    correct = clips[i]["text"].lower().strip()
    answer = message.text.lower().strip()
    state["attempted_text"] = True
    state["text_attempts"] += 1
    if answer == correct:
        state["text_attempts"] = 0
        await message.answer("✅ Верно! Теперь запиши аудио.")
    else:
        if state["text_attempts"] >= 2:
            keyboard = InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="👀 Показать ответ", callback_data="show_answer")]])
            await message.answer("❌ Ошибка. Попробуй ещё раз или нажми кнопку ниже.", reply_markup=keyboard)
        else:
            await message.answer("⚠️ Есть ошибки. Попробуй снова или запиши голосом.")

@dp.message(F.voice | F.audio)
async def handle_audio(message: types.Message):
    chat_id = message.chat.id
    state = user_state.get(chat_id)
    if not state:
        await message.answer("Начни с /start")
        return
    i = state["index"]
    voice = message.voice or message.audio
    file = await bot.get_file(voice.file_id)
    audio_path = os.path.join(USER_AUDIO_DIR, f"audio_{chat_id}_{i}.ogg")
    await bot.download_file(file.file_path, audio_path)
    recognized = transcribe(audio_path)
    score = compare(clips[i]["text"], recognized)
    await message.answer(f"🔊 Ты сказал: <i>{recognized}</i>\n📊 Произношение: <b>{score*100:.1f}%</b>")
    state["audio_sent"] = True
    if score <= 0.7:
        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔁 Ещё раз", callback_data="repeat_audio")],
            [InlineKeyboardButton(text="⏭ Продолжить", callback_data="force_next")]
        ])
        await message.answer("😕 Неточное произношение. Повторить или продолжить?", reply_markup=keyboard)
    else:
        await message.answer("👍 Хорошо! Перейди к следующему фрагменту:", reply_markup=continue_keyboard())

@dp.callback_query(lambda c: c.data == "repeat_audio")
async def repeat_audio(callback: CallbackQuery):
    await callback.message.answer("🎤 Жду повторную запись.")
    await callback.answer()

# --- Запуск ---
async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())